"""
Generates the synthetic sample data under data/ - entirely fictional schools,
names, and phone numbers (90000000xx range), so the repo is safe to publish
and the Quick Start works without any real data. Re-run any time to regenerate.
"""
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
DATA = BASE / "data"
DATA.mkdir(exist_ok=True)


def make_master():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Sample Master Contact List - Fictional District (all data synthetic)"
    ws.merge_cells("A1:C1")
    ws["A3"], ws["B3"], ws["C3"] = "Name of GSSS", "Name of Principal", "Mobile No"

    rows = [
        ("Rampur GSSS", "Anita Sharma", 9000000001),
        ("Sundarpur GSSS", "Rakesh Verma", 9000000002),
        ("Lakhanpur GSSS", "Meena Kumari", 9000000003),
        ("Chandpur GSSS", "Suresh Yadav", 9000000004),
        ("Devgarh GSSS", "Priya Singh", 9000000005),
        ("Kishangarh GSSS", "Vacant", None),  # vacant post -> no_contact_schools.csv demo
        ("Manoharpur GSSS", "Arvind Joshi", 9000000007),
        ("Fatehpur GSSS", "Sunita Rani", 9000000008),
        ("Govindpur GSSS", "Deepak Chand", 9000000009),
        ("Narsingpur GSSS", "Kavita Devi", 9000000010),
    ]
    r = 5
    for school, principal, mobile in rows:
        ws.cell(row=r, column=1, value=school)
        ws.cell(row=r, column=2, value=principal)
        ws.cell(row=r, column=3, value=mobile)
        r += 1
    wb.save(DATA / "sample_master_contacts.xlsx")
    print(f"Wrote {DATA / 'sample_master_contacts.xlsx'}")


def make_criteria_flat():
    import csv
    rows = [
        ["School", "Attendance Register", "Fee Receipt", "Inventory List", "Annual Report"],
        ["Rampur GSSS", "Yes", "Yes", "Yes", "Yes"],
        ["Sundarpur GSSS", "Yes", "", "Yes", "Yes"],
        ["Lakhanpur GSSS", "", "", "", ""],
        ["Chandpur GSSS", "Yes", "Yes", "", "Yes"],
        ["Kishangarh GSSS", "", "Yes", "Yes", ""],
        ["Rampur High School", "", "", "", ""],       # different school (GHS, not GSSS) - should NOT match Rampur GSSS
        ["Manohapur GSSS", "Yes", "Yes", "Yes", ""],  # typo of "Manoharpur GSSS" - fuzzy match demo
    ]
    path = DATA / "sample_criteria_flat.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)
    print(f"Wrote {path}")


def make_criteria_complex():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Sample Gender/Category-wise Statement for Class 9th & 10th (synthetic)"
    ws.merge_cells("A1:W1")
    ws["A2"], ws["B2"], ws["C2"], ws["D2"] = "Sr.No.", "Name of the School", "Class", "Categories"
    ws.merge_cells("D2:W2")
    groups = ["SC", "ST", "OBC", "BPL", "GEN"]
    start_col = 4
    for g in groups:
        ws.cell(row=3, column=start_col, value=g)
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=start_col + 3)
        for j, sub in enumerate(["B", "G", "T", "Remarks"]):
            ws.cell(row=4, column=start_col + j, value=sub)
        start_col += 4

    # (school, class, {group: (B,G,T) or None for fully missing, "partial" for one blank})
    data = [
        ("Rampur GSSS", "9th", {"SC": (5, 3, 8), "ST": (1, 0, 1), "OBC": (2, 2, 4), "BPL": (0, 0, 0), "GEN": (6, 4, 10)}),
        ("Rampur GSSS", "10th", {"SC": (4, 4, 8), "ST": (1, 1, 2), "OBC": (3, 1, 4), "BPL": (0, 0, 0), "GEN": (5, 5, 10)}),
        ("Sundarpur GSSS", "9th", {"SC": (2, 2, 4), "ST": None, "OBC": (1, 1, 2), "BPL": (0, 0, 0), "GEN": (3, 2, 5)}),
        ("Sundarpur GSSS", "10th", {"SC": (2, 3, 5), "ST": None, "OBC": (2, 0, 2), "BPL": (0, 0, 0), "GEN": (2, 2, 4)}),
        ("Lakhanpur GSSS", "9th", {"SC": None, "ST": None, "OBC": None, "BPL": None, "GEN": None}),
        ("Lakhanpur GSSS", "10th", {"SC": None, "ST": None, "OBC": None, "BPL": None, "GEN": None}),
        ("Devgarh GSSS", "9th", {"SC": (3, 3, 6), "ST": (0, 1, 1), "OBC": "partial", "BPL": (0, 0, 0), "GEN": (4, 4, 8)}),
        ("Devgarh GSSS", "10th", {"SC": (3, 2, 5), "ST": (1, 0, 1), "OBC": (5, 5, 10), "BPL": (0, 0, 0), "GEN": (2, 3, 5)}),
    ]

    r = 5
    sr = 1
    last_school = None
    for school, klass, groups_data in data:
        ws.cell(row=r, column=2, value=school if school != last_school else None)
        if school != last_school:
            ws.cell(row=r, column=1, value=sr)
            sr += 1
        last_school = school
        ws.cell(row=r, column=3, value=klass)
        start_col = 4
        for g in ["SC", "ST", "OBC", "BPL", "GEN"]:
            val = groups_data[g]
            if val is None:
                pass  # leave B/G/T/Remarks all blank
            elif val == "partial":
                ws.cell(row=r, column=start_col, value=2)      # B filled
                ws.cell(row=r, column=start_col + 1, value=None)  # G blank -> "incomplete"
                ws.cell(row=r, column=start_col + 2, value=2)   # T filled
            else:
                b, g, t = val
                ws.cell(row=r, column=start_col, value=b)
                ws.cell(row=r, column=start_col + 1, value=g)
                ws.cell(row=r, column=start_col + 2, value=t)
            start_col += 4
        r += 1

    wb.save(DATA / "sample_criteria_complex.xlsx")
    print(f"Wrote {DATA / 'sample_criteria_complex.xlsx'}")


if __name__ == "__main__":
    make_master()
    make_criteria_flat()
    make_criteria_complex()
