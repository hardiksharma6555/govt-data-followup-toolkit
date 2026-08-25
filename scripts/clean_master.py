"""
Stage 1: Clean a master contact list (school/office name, person, mobile) into
a canonical CSV used by every later stage.

Usage:
    python3 scripts/clean_master.py --input data/sample_master_contacts.xlsx

The header row doesn't have to be row 1 - real government sheets often have a
title row, a blank row, then the real header. This script scans the first 15
rows for cells that look like the three columns it needs (by keyword, not
fixed position) and starts reading data right after whichever row it finds.
If your file's headers don't match the default keywords, pass --school-col /
--principal-col / --mobile-col with the exact header text instead.
"""
import argparse
import csv
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import clean_mobile, normalize_name

SCHOOL_KEYWORDS = ["school", "gsss", "institution", "office", "centre", "center"]
PRINCIPAL_KEYWORDS = ["principal", "head", "name of principal", "contact person", "hm"]
MOBILE_KEYWORDS = ["mobile", "phone", "contact no", "whatsapp", "cell"]


def _matches(cell_value, keywords, explicit=None) -> bool:
    if not cell_value:
        return False
    text = str(cell_value).strip().lower()
    if explicit:
        return text == explicit.strip().lower()
    return any(k in text for k in keywords)


def find_header(ws, school_col_hint, principal_col_hint, mobile_col_hint, max_scan_rows=15):
    for r in range(1, max_scan_rows + 1):
        cols = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if _matches(val, SCHOOL_KEYWORDS, school_col_hint) and "school" not in cols:
                cols["school"] = c
            if _matches(val, PRINCIPAL_KEYWORDS, principal_col_hint) and "principal" not in cols:
                cols["principal"] = c
            if _matches(val, MOBILE_KEYWORDS, mobile_col_hint) and "mobile" not in cols:
                cols["mobile"] = c
        if "school" in cols and "mobile" in cols:
            return r, cols
    return None, None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, type=Path, help="Master contact list .xlsx")
    ap.add_argument("--output", type=Path, default=Path("output/master_clean.csv"))
    ap.add_argument("--sheet", default=None, help="Sheet name (default: first sheet)")
    ap.add_argument("--school-col", default=None, help="Exact header text for the school/office column, if auto-detect fails")
    ap.add_argument("--principal-col", default=None, help="Exact header text for the contact-person column")
    ap.add_argument("--mobile-col", default=None, help="Exact header text for the mobile number column")
    ap.add_argument("--country-code", default="91", help="Country code prefix for cleaned mobile numbers")
    ap.add_argument("--strip-tokens", default="GSSS,GSS", help="Comma-separated tokens to drop when normalizing school names")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]

    header_row, cols = find_header(ws, args.school_col, args.principal_col, args.mobile_col)
    if header_row is None:
        sys.exit(
            "Could not auto-detect the header row in the first 15 rows.\n"
            "Pass --school-col / --mobile-col with the exact header text from your file."
        )
    print(f"Detected header at row {header_row}: {cols}")

    strip_tokens = [t.strip() for t in args.strip_tokens.split(",") if t.strip()]

    # Skip any blank separator row(s) between the header and the first data
    # row (common in government sheets) instead of stopping immediately.
    r = header_row + 1
    lookahead_limit = r + 10
    while r < lookahead_limit:
        if ws.cell(row=r, column=cols["school"]).value or ws.cell(row=r, column=cols["mobile"]).value:
            break
        r += 1

    rows, bad_mobile = [], []
    while True:
        school = ws.cell(row=r, column=cols["school"]).value
        principal = ws.cell(row=r, column=cols.get("principal", cols["school"])).value if "principal" in cols else None
        mobile_raw = ws.cell(row=r, column=cols["mobile"]).value
        if school is None and mobile_raw is None:
            break
        school = (school or "").strip() if isinstance(school, str) else (str(school).strip() if school else "")
        principal = (principal or "").strip() if isinstance(principal, str) else (str(principal).strip() if principal else "")
        mobile = clean_mobile(mobile_raw, args.country_code)
        if len(mobile) != len(args.country_code) + 10:
            bad_mobile.append((school, principal, mobile_raw))
        rows.append({
            "school_raw": school,
            "school_norm": normalize_name(school, strip_tokens),
            "principal_name": principal,
            "mobile": mobile,
        })
        r += 1

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["school_raw", "school_norm", "principal_name", "mobile"])
        w.writeheader()
        w.writerows(rows)

    print(f"Wrote {len(rows)} contact rows -> {args.output}")
    if bad_mobile:
        print(f"WARNING: {len(bad_mobile)} rows have a mobile number that isn't a clean 10-digit number (vacant posts, typos, etc.) - these will be skipped at send time:")
        for school, principal, raw in bad_mobile:
            print(f"  - {school!r} / {principal!r}: {raw!r}")


if __name__ == "__main__":
    main()
