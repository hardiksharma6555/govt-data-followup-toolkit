"""
Stage 2 (generic path): Analyze a FLAT criteria file - one row per school/office,
first column is the identifier, every other column is a field that should be
filled in. Any blank cell in a data column is reported as missing.

Works for both .csv and .xlsx. This covers most "secondary CSV" cases: a
list of schools with a handful of yes/no or numeric columns where empty means
"not submitted yet".

If your criteria file has a complex multi-row header (merged category groups,
sub-columns, etc. - like a government FTB/attendance return) this generic
reader won't understand the structure. Copy scripts/analyze_complex_example.py
instead and adapt it - see README "Writing a custom analyzer".

Usage:
    python3 scripts/analyze_generic.py --input data/sample_criteria_flat.csv
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import normalize_name


def read_rows(path: Path, sheet: str | None):
    if path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            return list(reader)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    return [[c.value for c in row] for row in ws.iter_rows()]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, type=Path, help="Criteria file (.csv or .xlsx)")
    ap.add_argument("--output", type=Path, default=Path("output/missing_report.csv"))
    ap.add_argument("--sheet", default=None, help="Sheet name, for .xlsx input (default: first sheet)")
    ap.add_argument("--strip-tokens", default="GSSS,GSS", help="Comma-separated tokens to drop when normalizing the identifier column")
    args = ap.parse_args()

    raw_rows = [r for r in read_rows(args.input, args.sheet) if any(c not in (None, "") for c in r)]
    if len(raw_rows) < 2:
        sys.exit("File needs a header row plus at least one data row.")

    header, data_rows = raw_rows[0], raw_rows[1:]
    field_names = header[1:]
    strip_tokens = [t.strip() for t in args.strip_tokens.split(",") if t.strip()]

    out_rows = []
    for row in data_rows:
        identifier = str(row[0]).strip() if row[0] not in (None, "") else ""
        if not identifier:
            continue
        missing = []
        for i, field in enumerate(field_names, start=1):
            val = row[i] if i < len(row) else None
            if val in (None, ""):
                missing.append(str(field).strip())
        out_rows.append({
            "school_raw": identifier,
            "school_norm": normalize_name(identifier, strip_tokens),
            "class": "",
            "missing_categories": "; ".join(missing),
            "incomplete_categories": "",
        })

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["school_raw", "school_norm", "class", "missing_categories", "incomplete_categories"])
        w.writeheader()
        w.writerows(out_rows)

    flagged = [r for r in out_rows if r["missing_categories"]]
    print(f"Wrote {len(out_rows)} rows -> {args.output}")
    print(f"{len(flagged)} rows have at least one missing field:")
    for r in flagged:
        print(f"  - {r['school_raw']!r}: MISSING {r['missing_categories']}")


if __name__ == "__main__":
    main()
