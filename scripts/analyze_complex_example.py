"""
Stage 2 (complex-layout EXAMPLE): analyzer for a criteria file with a
multi-row merged header - category groups split into sub-columns - which is
common in Indian government returns (e.g. "Gender/Category-wise Free Text
Books statement": SC/ST/OBC/BPL/GEN groups, each split into Boys/Girls/
Total/Remarks).

This file is a REFERENCE TEMPLATE, not a general-purpose tool: multi-row
header layouts vary too much between departments/forms to auto-detect
reliably. Copy this file, rename it, and adjust CATEGORY_GROUPS + the row
offsets to match your actual sheet before running it - see README "Writing a
custom analyzer".

Demonstrated here against data/sample_criteria_complex.xlsx, which mirrors
the real-world layout with synthetic schools/numbers:
  row 1: title (merged)
  row 2: Sr.No | School | Class | Categories (merged over all category cols)
  row 3: category group label (SC / ST / OBC / BPL / GEN), each spanning 4 cols
  row 4: sub-label under each group: B / G / T / Remarks
  row 5+: data, one row per (school, class)

A category/class is "missing" if ALL of its B/G/T sub-columns are blank, and
"incomplete" if only some are blank. The 4th sub-column (Remarks) is excluded
from the check by default - see README for why treating a free-text remarks
column as a required field usually produces false positives.
"""
import argparse
import csv
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from common import normalize_name

# (group label, [B, G, T] column indices) - EDIT to match your sheet.
# Column 4 of each group (Remarks) is intentionally left out of the check.
CATEGORY_GROUPS = [
    ("SC", [4, 5, 6]),
    ("ST", [8, 9, 10]),
    ("OBC", [12, 13, 14]),
    ("BPL", [16, 17, 18]),
    ("GEN", [20, 21, 22]),
]
DATA_START_ROW = 5
SCHOOL_COL = 2
CLASS_COL = 3


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--output", type=Path, default=Path("output/missing_report.csv"))
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--strip-tokens", default="GSSS,GSS")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    strip_tokens = [t.strip() for t in args.strip_tokens.split(",") if t.strip()]

    rows = []
    current_school = None
    r = DATA_START_ROW
    while True:
        school_cell = ws.cell(row=r, column=SCHOOL_COL).value
        klass = ws.cell(row=r, column=CLASS_COL).value
        if school_cell is None and klass is None:
            break
        if school_cell:
            current_school = str(school_cell).strip()

        missing, incomplete = [], []
        for label, cols in CATEGORY_GROUPS:
            vals = [ws.cell(row=r, column=c).value for c in cols]
            blanks = sum(1 for v in vals if v is None)
            if blanks == len(vals):
                missing.append(label)
            elif blanks > 0:
                incomplete.append(label)

        rows.append({
            "school_raw": current_school,
            "school_norm": normalize_name(current_school, strip_tokens),
            "class": klass,
            "missing_categories": "; ".join(missing),
            "incomplete_categories": "; ".join(incomplete),
        })
        r += 1

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["school_raw", "school_norm", "class", "missing_categories", "incomplete_categories"])
        w.writeheader()
        w.writerows(rows)

    flagged = [r for r in rows if r["missing_categories"] or r["incomplete_categories"]]
    print(f"Wrote {len(rows)} school/class rows -> {args.output}")
    print(f"{len(flagged)} rows have at least one missing or incomplete category:")
    for r in flagged:
        bits = []
        if r["missing_categories"]:
            bits.append(f"MISSING: {r['missing_categories']}")
        if r["incomplete_categories"]:
            bits.append(f"INCOMPLETE: {r['incomplete_categories']}")
        print(f"  - {r['school_raw']!r} ({r['class']}): {' | '.join(bits)}")


if __name__ == "__main__":
    main()
